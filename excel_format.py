import openpyxl


def get_device_id():
    wb = openpyxl.load_workbook('import.xlsx')
    sheet1 = wb['sheet1']
    nrows = sheet1.max_row + 1
    for j in range(1, nrows):
        value = sheet1.cell(j, 2).value
        if value == 0:
            print('剩余码数为：', sheet1.max_row - j)
            return sheet1.cell(j, 1).value, j


def set_token_used(k):
    wb = openpyxl.load_workbook("import.xlsx")
    sheet1 = wb['sheet1']
    sheet1.cell(k, 2).value = "1"
    wb.save("import.xlsx")
    pass
