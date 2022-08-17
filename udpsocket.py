import os
import socket
import threading
import time

import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfilename

import udp

file: str
udp_flag: bool = False
flow_id: int
left_num: int
udp: socket


def get_now_time():
    now = time.localtime()
    now_time = time.strftime("%Y%m%d", now)
    return now_time


def get_device_id():
    print('filename is ', str(file))
    wb = openpyxl.load_workbook(str(file))
    sheet1 = wb['sheet1']
    nrows = sheet1.max_row + 1
    for j in range(1, nrows):
        value = sheet1.cell(j, 2).value
        global flow_id
        global left_num
        left_num = sheet1.max_row - j
        flow_id = j - 1
        if value == 0:
            print('剩余码数为：', left_num)
            return sheet1.cell(j, 1).value, j


def set_token_used(k):
    wb = openpyxl.load_workbook(str(file))
    sheet1 = wb['sheet1']
    sheet1.cell(k, 2).value = "1"
    wb.save(str(file))
    pass


def display_off():
    string = ''
    result.set(string)


def display_result():
    global flow_id
    string = '流水号', flow_id, "烧录成功！ 下一台..."
    result.set(string)
    root.after(3000, display_off())


def start_udp():
    global udp
    udp = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    print('start udp server...\n')

    print(socket.gethostbyname(socket.getfqdn(socket.gethostname())))
    try:
        udp.bind((socket.gethostbyname(socket.getfqdn(socket.gethostname())), 1777))
    except:
        pass
    global udp_flag
    udp_flag = True

    while udp_flag:
        rec_msg, addr = udp.recvfrom(1024)
        client_ip, client_port = addr
        print('client_ip:', client_ip, 'client_port:', client_port)

        print('msg from client:', rec_msg.decode('utf8'))
        if rec_msg.decode('utf8') == "recv":
            try:
                j = get_device_id()[1]
                set_token_used(j)
                print('烧录成功！\n')
                display_result()
            except:
                print('无可用序列号，请重新导入！！\n')
                pass

        if rec_msg.decode('utf8') == "ask1":
            try:
                ack_msg = get_device_id()[0]
                print(ack_msg)
                udp.sendto(ack_msg.encode('utf8'), addr)
            except:
                print('无可用序列号，请重新导入！！\n')
                pass

        if rec_msg.decode('utf8') == "ask2":
            ack_msg = get_now_time()
            print(ack_msg)
            udp.sendto(ack_msg.encode('utf8'), addr)

        if rec_msg.decode('utf8') == "ask3":
            try:
                ack_msg = get_device_id()[1] - 1
                ack_msg = str(ack_msg)
                print(ack_msg)
                udp.sendto(ack_msg.encode('utf8'), addr)
            except:
                print('无可用序列号，请重新导入！！\n')
                pass

        if not udp_flag:
            print('stop udp server\n')
            break


def stop_udp():
    global udp_flag
    udp_flag = False
    global udp
    udp.sendto('123'.encode('utf8'), (socket.gethostbyname(socket.getfqdn(socket.gethostname())), 1777))


def select_open_file():
    global file
    file = askopenfilename(title='选择文件', filetypes=[('excel文件', '*.xlsx')])
    print('file:', file)
    get_device_id()
    get_flow_id()


def start_udp_thread():
    udp_thread = threading.Thread(target=start_udp, name='aa')
    udp_thread.start()


def get_flow_id():
    global flow_id
    global left_num
    global udp_flag

    string = '流水号：', flow_id, '剩余码数：', left_num, '运行状态：', '运行' if udp_flag else '停止'
    dstr.set(string)
    root.after(3000, get_flow_id)


root = tk.Tk()
root.geometry("400x360")
root.title("ipc烧录工具")
file = './import.xlsx'
longtext = """
操作指南：

  1.选择导入设备id的excel文件。
  2.点击开始运行
  3.自动烧录
"""
tk.Label(root, text=longtext, anchor="w", justify="left").pack()

tk.Button(root, text='选择文件', width=15, height=2, command=select_open_file).pack()
tk.Button(root, text="开始运行", width=15, height=2, command=start_udp_thread).pack()
tk.Button(root, text="停止运行", width=15, height=2, command=stop_udp).pack()

dstr = tk.StringVar()
lb = tk.Label(root, textvariable=dstr)
lb.pack()

result = tk.StringVar()
lb2 = tk.Label(root, textvariable=result, width=30, height=2, fg='red')
lb2.pack()


root.mainloop()
